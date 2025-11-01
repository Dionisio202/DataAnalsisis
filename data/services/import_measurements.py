import csv, io
from datetime import datetime
from typing import Any, Optional
from openpyxl import load_workbook  # pip install openpyxl
from ..models import Measurement, Ambiente

# Ajustes
BATCH_SIZE = 5000
ROUND_DECIMALS = 1  # redondeo para temp/noise/humidity (ajusta o pon None para desactivar)

# formatos de fecha aceptados (CSV o Excel como texto)
DATETIME_FORMATS = [
    "%m/%d/%Y %H:%M",
    "%m/%d/%Y %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d %H:%M:%S",
]

def _parse_datetime_flexible(v: Any) -> datetime:
    """Acepta datetime nativo, string en varios formatos o lanza ValueError amigable."""
    if v is None or v == "":
        raise ValueError("Fecha vacía")
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in DATETIME_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f"Formato de fecha no reconocido: '{s}'")

def _to_float(v: Any, field: str) -> float:
    if v is None or v == "":
        raise ValueError(f"{field} vacío")
    val = float(str(v).strip())
    if ROUND_DECIMALS is not None:
        val = round(val, ROUND_DECIMALS)
    return val

def _to_int(v: Any, field: str) -> int:
    if v is None or v == "":
        raise ValueError(f"{field} vacío")
    # acepta strings y números con .0
    return int(float(str(v).strip()))

def _clean_ambiente(s: Optional[str]) -> str:
    s = (s or "").strip().lower()
    if "óptimo" in s or "optimo" in s:
        return Ambiente.OPTIMO
    if "acept" in s:
        return Ambiente.ACEPTABLE
    return Ambiente.CRITICO

def _row_is_totally_empty(d: dict) -> bool:
    return all((v is None or str(v).strip() == "") for v in d.values())

def import_measurements_file(django_file):
    """
    Recibe un archivo subido (CSV/XLSX) y devuelve (inserted, errors:list[str]).
    Cabeceras esperadas (sin importar mayúsculas/espacios/BOM):
      creat_at, valtem, valco2, valson, valhum, ambiente
      opcionales: tvoc, weight
    """
    f = django_file
    created, errors = 0, []

    # ---------------- CSV ----------------
    if f.name.lower().endswith(".csv"):
        text = io.TextIOWrapper(f.file, encoding="utf-8", errors="ignore")
        reader = csv.DictReader(text)

        # normaliza headers: minus, strip, quita BOM
        norm_headers = [
            (h or "").strip().lower().lstrip("\ufeff")
            for h in (reader.fieldnames or [])
        ]
        reader.fieldnames = norm_headers

        # alias para fecha
        creat_key = "creat_at" if "creat_at" in norm_headers else (
            "created_at" if "created_at" in norm_headers else None
        )
        if not creat_key:
            return 0, ["No se encontró la columna de fecha (creat_at/created_at)."]

        batch = []
        for i, row in enumerate(reader, start=2):  # fila de datos real
            try:
                # dict con claves normalizadas a minúscula
                # (DictReader ya usa fieldnames normalizados arriba)
                if _row_is_totally_empty(row):
                    continue

                m = Measurement(
                    created_at=_parse_datetime_flexible(row.get(creat_key)),
                    temp_c=_to_float(row.get("valtem"), "valTem"),
                    co2_ppm=_to_int(row.get("valco2"), "valCO2"),
                    noise_dba=_to_float(row.get("valson"), "valSon"),
                    humidity_pct=_to_float(row.get("valhum"), "valHum"),
                    ambiente=_clean_ambiente(row.get("ambiente")),
                    tvoc_index=_to_int(row.get("tvoc"), "tvoc") if row.get("tvoc") not in (None, "") else None,
                    weight=_to_float(row.get("weight"), "weight") if row.get("weight") not in (None, "") else None,
                )
                batch.append(m)
                if len(batch) >= BATCH_SIZE:
                    Measurement.objects.bulk_create(batch, ignore_conflicts=True)
                    created += len(batch)
                    batch.clear()
            except Exception as e:
                errors.append(f"Fila {i}: {e}")

        if batch:
            Measurement.objects.bulk_create(batch, ignore_conflicts=True)
            created += len(batch)

        return created, errors

    # ---------------- XLSX / XLSM ----------------
    elif f.name.lower().endswith((".xlsx", ".xlsm")):
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active

        # normaliza headers a minúsculas
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [
            (str(c.value).strip().lower() if c.value is not None else "")
            for c in header_row
        ]
        idx = {h: i for i, h in enumerate(headers)}

        # alias para fecha
        if "creat_at" in idx:
            creat_key = "creat_at"
        elif "created_at" in idx:
            creat_key = "created_at"
        else:
            return 0, ["Columna requerida faltante: creat_at (o created_at)"]

        # valida columnas mínimas
        required = ["valtem", "valco2", "valson", "valhum", "ambiente"]
        missing = [r for r in required if r not in idx]
        if missing:
            return 0, [f"Columnas requeridas faltantes: {', '.join(missing)}"]

        def cell(row, name):
            return row[idx[name]].value if name in idx else None

        batch = []
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                row_dict = {
                    "fecha": cell(row, creat_key),
                    "valtem": cell(row, "valtem"),
                    "valco2": cell(row, "valco2"),
                    "valson": cell(row, "valson"),
                    "valhum": cell(row, "valhum"),
                    "ambiente": cell(row, "ambiente"),
                    "tvoc": cell(row, "tvoc") if "tvoc" in idx else None,
                    "weight": cell(row, "weight") if "weight" in idx else None,
                }
                if _row_is_totally_empty(row_dict):
                    continue

                m = Measurement(
                    created_at=_parse_datetime_flexible(row_dict["fecha"]),
                    temp_c=_to_float(row_dict["valtem"], "valTem"),
                    co2_ppm=_to_int(row_dict["valco2"], "valCO2"),
                    noise_dba=_to_float(row_dict["valson"], "valSon"),
                    humidity_pct=_to_float(row_dict["valhum"], "valHum"),
                    ambiente=_clean_ambiente(row_dict["ambiente"]),
                    tvoc_index=_to_int(row_dict["tvoc"], "tvoc") if row_dict["tvoc"] not in (None, "") else None,
                    weight=_to_float(row_dict["weight"], "weight") if row_dict["weight"] not in (None, "") else None,
                )
                batch.append(m)
                if len(batch) >= BATCH_SIZE:
                    Measurement.objects.bulk_create(batch, ignore_conflicts=True)
                    created += len(batch)
                    batch.clear()
            except Exception as e:
                errors.append(f"Fila {i}: {e}")

        if batch:
            Measurement.objects.bulk_create(batch, ignore_conflicts=True)
            created += len(batch)

        return created, errors

    # ---------------- Otros ----------------
    else:
        raise ValueError("Formato no soportado. Usa .csv o .xlsx")
